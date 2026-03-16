"""
campaign.py — Campaign orchestrator.

Ties together ingestion, draft generation, review, and sending
into a coherent workflow. This is the main interface used by outreach.py.
"""

import logging
from pathlib import Path

from . import outreach_config as cfg
from .state import OutreachDB
from .safety import validate_email, check_from_address
from .email_generator import generate_drafts_batch
from .sender import send_batch

logger = logging.getLogger("outreach")


# ---------------------------------------------------------------------------
# Excel / CSV ingestion
# ---------------------------------------------------------------------------
def _load_leads_from_excel(path: str) -> list[dict]:
    """
    Read leads from the existing lead_tracker.xlsx file.

    This reads the Excel file produced by your existing pipeline
    and converts each row into a dict suitable for the outreach DB.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True)
    ws = wb.active

    # Read headers from row 1
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value else "")

    # Map Excel column labels to our field names
    label_map = {
        "Business Name": "business_name",
        "Phone": "phone",
        "Category": "primary_category",
        "City": "city",
        "Rating": "rating",
        "Reviews": "review_count",
        "Lead Score": "lead_score",
        "Website Status": "website_status",
        "Email": "email",
        "Notes": "notes",
        "Contacted": "contacted",
    }

    leads = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        lead = {}
        for cell, header in zip(row, headers):
            field = label_map.get(header)
            if field:
                lead[field] = cell.value if cell.value is not None else ""
        # Only include rows that have a business name and email
        if lead.get("business_name") and lead.get("email"):
            leads.append(lead)

    wb.close()
    return leads


def _load_leads_from_csv(path: str) -> list[dict]:
    """Read leads from a CSV file."""
    import csv

    leads = []
    encodings = ["utf-8-sig", "utf-8", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            with open(path, encoding=enc) as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row.get("email") or row.get("Email"):
                        lead = {
                            "business_name": row.get("business_name", row.get("Business Name", "")),
                            "email": row.get("email", row.get("Email", "")),
                            "website": row.get("website", row.get("Website", "")),
                            "city": row.get("city", row.get("City", "")),
                            "primary_category": row.get("category", row.get("Category", "")),
                            "lead_score": row.get("lead_score", row.get("Lead Score", 0)),
                            "phone": row.get("phone", row.get("Phone", "")),
                            "rating": row.get("rating", row.get("Rating", 0)),
                            "review_count": row.get("review_count", row.get("Reviews", 0)),
                        }
                        leads.append(lead)
            break
        except UnicodeDecodeError:
            continue

    return leads


def ingest_leads(file_path: str = "") -> tuple[int, int]:
    """
    Load leads from Excel or CSV and import into the outreach database.

    Returns (inserted, skipped).
    """
    path = file_path or cfg.LEAD_EXCEL_PATH
    if not Path(path).exists():
        raise FileNotFoundError(f"Lead file not found: {path}")

    # Detect format
    if path.endswith(".csv"):
        leads = _load_leads_from_csv(path)
    else:
        leads = _load_leads_from_excel(path)

    if not leads:
        logger.warning("No leads with emails found in %s", path)
        return 0, 0

    # Filter out invalid emails before ingestion
    valid_leads = []
    invalid_count = 0
    for lead in leads:
        ok, reason = validate_email(lead.get("email", ""))
        if ok:
            valid_leads.append(lead)
        else:
            logger.debug("Skipping %s: %s", lead.get("email", "?"), reason)
            invalid_count += 1

    if invalid_count:
        logger.info("Filtered out %d leads with invalid emails", invalid_count)

    db = OutreachDB()
    try:
        inserted, skipped = db.ingest_many(valid_leads)
    finally:
        db.close()

    return inserted, skipped


# ---------------------------------------------------------------------------
# Draft generation
# ---------------------------------------------------------------------------
def generate_all_drafts() -> tuple[int, int]:
    """
    Generate email drafts for all new leads above the score threshold.

    Returns (success_count, error_count).
    """
    if not cfg.ANTHROPIC_API_KEY:
        raise RuntimeError("ANTHROPIC_API_KEY not set — cannot generate drafts")

    db = OutreachDB()
    try:
        new_leads = db.get_new_leads(min_score=cfg.MIN_SCORE_THRESHOLD)
        if not new_leads:
            logger.info("No new leads above score threshold %d", cfg.MIN_SCORE_THRESHOLD)
            return 0, 0

        logger.info("Generating drafts for %d leads (score >= %d)",
                    len(new_leads), cfg.MIN_SCORE_THRESHOLD)
        return generate_drafts_batch(new_leads, db)
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Review workflow
# ---------------------------------------------------------------------------
def get_review_queue() -> list[dict]:
    """Get leads with drafts ready for human review."""
    db = OutreachDB()
    try:
        return db.get_leads_needing_review()
    finally:
        db.close()


def approve_lead(email: str, notes: str = ""):
    """Approve a lead for sending."""
    db = OutreachDB()
    try:
        db.mark_approved(email, notes)
    finally:
        db.close()


def reject_lead(email: str, notes: str = ""):
    """Reject a lead."""
    db = OutreachDB()
    try:
        db.mark_rejected(email, notes)
    finally:
        db.close()


def get_leads_needing_review() -> list[dict]:
    """Get leads with generated drafts that need human review."""
    db = OutreachDB()
    try:
        return db.get_leads_needing_review()
    finally:
        db.close()


def approve_all_reviewed():
    """Approve all leads currently in Reviewed status."""
    db = OutreachDB()
    try:
        reviewed = db.get_leads_needing_review()
        count = 0
        for lead in reviewed:
            db.mark_approved(lead["email"])
            count += 1
        return count
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Sending
# ---------------------------------------------------------------------------
def send_approved(dry_run: bool = False) -> tuple[int, int, int]:
    """
    Send emails to all approved leads.

    Args:
        dry_run: If True, simulate sends without actually sending.

    Returns:
        (sent_count, failed_count, skipped_count)
    """
    # Pre-flight checks
    if not dry_run:
        ok, reason = check_from_address()
        if not ok:
            raise RuntimeError(f"Cannot send: {reason}")

    db = OutreachDB()
    try:
        leads = db.get_sendable_leads()
        if not leads:
            logger.info("No approved leads ready to send")
            return 0, 0, 0

        logger.info("Found %d approved leads to send%s",
                    len(leads), " (DRY RUN)" if dry_run else "")
        return send_batch(leads, db, dry_run=dry_run)
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Status and stats
# ---------------------------------------------------------------------------
def get_campaign_stats() -> dict:
    """Get campaign statistics."""
    db = OutreachDB()
    try:
        return db.get_stats()
    finally:
        db.close()


def get_all_leads() -> list[dict]:
    """Get all leads in the database."""
    db = OutreachDB()
    try:
        return db.get_all_leads()
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Opt-out management
# ---------------------------------------------------------------------------
def add_opt_out(email: str, reason: str = ""):
    """Add an email to the opt-out list and mark as DoNotContact."""
    db = OutreachDB()
    try:
        db.mark_do_not_contact(email, reason)
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Full pipeline (ingest → draft → review prompt)
# ---------------------------------------------------------------------------
def run_ingest_pipeline(file_path: str = "") -> dict:
    """
    Run the full ingestion pipeline:
    1. Load leads from Excel/CSV
    2. Import to database (dedup)
    3. Generate drafts for new leads

    Returns a summary dict.
    """
    summary = {}

    # Step 1: Ingest
    inserted, skipped = ingest_leads(file_path)
    summary["ingested"] = inserted
    summary["skipped_duplicates"] = skipped
    logger.info("Ingested %d new leads (%d duplicates skipped)", inserted, skipped)

    # Step 2: Generate drafts
    if cfg.ANTHROPIC_API_KEY:
        success, errors = generate_all_drafts()
        summary["drafts_generated"] = success
        summary["draft_errors"] = errors
        logger.info("Generated %d drafts (%d errors)", success, errors)
    else:
        summary["drafts_generated"] = 0
        summary["draft_errors"] = 0
        if not cfg.ANTHROPIC_API_KEY:
            logger.warning("ANTHROPIC_API_KEY not set — skipping draft generation")

    return summary
