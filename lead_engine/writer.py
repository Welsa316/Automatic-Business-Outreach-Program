"""
writer.py — Excel output with append mode.

Produces a single Excel tracker file. On re-runs, existing rows
(and user edits like Contacted status and Notes) are preserved;
only new businesses are appended.
"""

import logging
from pathlib import Path

from . import config

logger = logging.getLogger("lead_engine")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# Excel columns: (field_key, header_label, width)
EXCEL_COLUMNS = [
    ("contacted",        "Contacted",      12),
    ("business_name",    "Business Name",  30),
    ("phone",            "Phone",          16),
    ("primary_category", "Category",       20),
    ("city",             "City",           16),
    ("rating",           "Rating",          8),
    ("review_count",     "Reviews",        10),
    ("lead_score",       "Lead Score",     11),
    ("has_website",      "Has Website",    13),
    ("notes",            "Notes",          30),
]


def _dedup_key(name: str, phone: str) -> str:
    """Create a dedup key from business name + phone."""
    return f"{name.strip().lower()}|{phone.strip()}"


def _biz_to_row(biz: dict) -> dict:
    """Flatten a business dict into the output columns."""
    return {
        "contacted":        biz.get("contacted", ""),
        "business_name":    biz.get("business_name", ""),
        "phone":            biz.get("phone", ""),
        "primary_category": biz.get("primary_category", ""),
        "city":             biz.get("city", ""),
        "rating":           biz.get("rating", ""),
        "review_count":     biz.get("review_count", ""),
        "lead_score":       biz.get("lead_score", 0),
        "has_website":      "Yes" if biz.get("has_website") else "No",
        "notes":            biz.get("notes", ""),
    }


def _load_existing(path: Path) -> list[dict]:
    """
    Load existing rows from an Excel tracker file.
    Returns a list of row dicts with the same keys as EXCEL_COLUMNS.
    """
    if not path.exists() or not _HAS_OPENPYXL:
        return []

    try:
        wb = load_workbook(str(path))
        ws = wb.active
    except Exception as exc:
        logger.warning("Could not read existing Excel file: %s", exc)
        return []

    # Read headers from row 1 to map columns
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        headers.append(str(val).strip() if val else "")

    # Build reverse map: header label → field key
    label_to_key = {label: key for key, label, _ in EXCEL_COLUMNS}

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            field_key = label_to_key.get(header)
            if field_key:
                val = ws.cell(row=row_idx, column=col_idx).value
                row_data[field_key] = val if val is not None else ""
        # Only include rows that have a business name
        if row_data.get("business_name"):
            rows.append(row_data)

    wb.close()
    logger.info("Loaded %d existing rows from %s", len(rows), path.name)
    return rows


def _merge_rows(existing: list[dict], new_businesses: list[dict]) -> list[dict]:
    """
    Merge existing Excel rows with new business data.
    - Existing rows are kept as-is (preserving Contacted, Notes, etc.)
    - New businesses not already in the Excel are appended
    - Result is sorted by lead_score descending
    """
    # Index existing rows by dedup key
    seen = {}
    merged = []
    for row in existing:
        key = _dedup_key(row.get("business_name", ""), str(row.get("phone", "")))
        if key not in seen:
            seen[key] = True
            merged.append(row)

    # Add new businesses that aren't already tracked
    added = 0
    for biz in new_businesses:
        key = _dedup_key(biz.get("business_name", ""), str(biz.get("phone", "")))
        if key not in seen:
            seen[key] = True
            merged.append(_biz_to_row(biz))
            added += 1

    # Sort by lead_score descending
    def _score(row):
        try:
            return int(row.get("lead_score", 0))
        except (ValueError, TypeError):
            return 0

    merged.sort(key=_score, reverse=True)

    logger.info("Merge: %d existing + %d new = %d total",
                len(existing), added, len(merged))
    return merged


def _build_sheet(ws, rows: list[dict]) -> None:
    """Populate a worksheet with lead data and formatting."""
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4A4A8A", end_color="4A4A8A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

    # Write headers
    for col_idx, (key, label, width) in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, (key, label, width) in enumerate(EXCEL_COLUMNS, 1):
            value = row_data.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    total_rows = len(rows) + 1  # +1 for header

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    if total_rows > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLUMNS))}{total_rows}"

    # Data validation for Contacted column (Yes/No dropdown)
    contacted_col = 1  # first column
    if total_rows > 1:
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        dv.error = "Please select Yes or No"
        dv.errorTitle = "Invalid Entry"
        col_letter = get_column_letter(contacted_col)
        dv.sqref = f"{col_letter}2:{col_letter}{total_rows}"
        ws.add_data_validation(dv)

    # Conditional formatting: Contacted = "Yes" → green
    if total_rows > 1:
        col_letter = get_column_letter(contacted_col)
        contacted_range = f"{col_letter}2:{col_letter}{total_rows}"
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ws.conditional_formatting.add(contacted_range, CellIsRule(
            operator="equal",
            formula=['"Yes"'],
            fill=green_fill,
        ))

    # Conditional formatting for Lead Score column
    score_col = None
    for col_idx, (key, label, width) in enumerate(EXCEL_COLUMNS, 1):
        if key == "lead_score":
            score_col = col_idx
            break

    if score_col and total_rows > 1:
        col_letter = get_column_letter(score_col)
        score_range = f"{col_letter}2:{col_letter}{total_rows}"

        # Green for high scores
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(color="006100", bold=True)
        ws.conditional_formatting.add(score_range, CellIsRule(
            operator="greaterThanOrEqual",
            formula=[str(config.EXCEL_HIGH_SCORE_THRESHOLD)],
            fill=green_fill, font=green_font,
        ))

        # Yellow for medium scores
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        yellow_font = Font(color="9C6500")
        ws.conditional_formatting.add(score_range, CellIsRule(
            operator="between",
            formula=[str(config.EXCEL_MEDIUM_SCORE_THRESHOLD),
                     str(config.EXCEL_HIGH_SCORE_THRESHOLD - 1)],
            fill=yellow_fill, font=yellow_font,
        ))


def write_outputs(businesses: list[dict], output_dir: str | Path) -> dict:
    """
    Write Excel tracker with append mode.

    If the Excel file already exists, existing rows (and user edits)
    are preserved. Only new businesses are added.

    Returns a dict of {file_type: Path}.
    """
    if not _HAS_OPENPYXL:
        logger.error("openpyxl not installed — cannot create Excel output. "
                     "Install with: pip install openpyxl")
        return {}

    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    excel_path = out / config.EXCEL_FILENAME

    # Load existing data (if any)
    existing = _load_existing(excel_path)

    # Merge with new businesses
    merged = _merge_rows(existing, businesses)

    # Write fresh Excel with all rows
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    _build_sheet(ws, merged)

    wb.save(str(excel_path))
    logger.info("Saved Excel → %s (%d leads)", excel_path, len(merged))

    return {"excel": excel_path}
